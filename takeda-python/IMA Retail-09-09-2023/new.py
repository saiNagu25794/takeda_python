import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import sys
from Auto_preview import auto_review_dict


def process_input_file(input_file, output_file):
    #reload the workbook
    workbook = load_workbook(output_file)
    sheet = workbook['Auto']

    input_df = pd.read_excel(input_file)

    for index, row in input_df.iterrows():
        file_name = row['File']
        col_index = sheet.max_column + 1
        column_header = sheet.cell(row=1, column=col_index, value=file_name)
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = 45.00
        column_header.fill= PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
        for key in auto_review_dict.keys():
            if key in input_df.columns:
                mapped_key = key
            else:
                mapped_key = auto_review_dict[key]

            sheet.cell(row=list(auto_review_dict.keys()).index(key) + 2, column = col_index, value = row[mapped_key])


    # Enable text wrapping for all rows
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')



    workbook.save(output_file)

if __name__ == "__main__":


    input_file = r"C:\Users\rrguest.RR-ITS\Documents\IMA Retail-09-09-2023\IMA Retail\Input Excels\Auto_preview.xlsx"
    output_file = r'C:\Users\rrguest.RR-ITS\Documents\IMA Retail-09-09-2023\IMA Retail\output.xlsx'

    process_input_file(input_file, output_file)