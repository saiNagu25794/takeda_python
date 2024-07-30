import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

auto_review_dict = {
    'Carrier': "",
    "Named Insured": "Insured Name",
    'Liability': "",
    'Uninsured/Underinsured Motorists Per Accident': "Uninsured or Underinsured Motorists Per Accident",
    'Comprehensive': "",
    'Collision': "",
    'Medical Payments Per Person': "",
    "Scheduled Auto's Comprehensive": "Scheduled Autos Comprehensive",
    "Scheduled Auto's Collision": "Scheduled Autos Collision",
    'Garagekeepers Coverage Each customer': "",
    'Broadened Who Is An Insured': "",
    'Blanket Additional Insured': "",
    'Liability Coverage Extensions Supplementary Payments - Increased Limits': "",
    'Newly Acquired or Formed Organizations': "",
    'Extended Towing': "",
    'Physical Damage Coverage Extension - Transportation Expenses': "",
    'Physical Damage Coverage Extension - Loss of Use': "",
    'Rental Reimbursement': "",
    'Airbags': "",
    'Audio, Visual, and Data Electronic Equipment': "Audio Visual and Data Electronic Equipment",
    'Loan/Lease Payoff Coverage': "Loan or Lease Payoff Coverage",
    'Hired Auto Physical Damage': "",
    'Waiver of Subrogation': "",
    'Extended Employee Hired Auto Physical Damage': "",
    'Fellow Employee Exclusion Removed': "",
    'Definition of Bodily Injury Amended': "",
    '2022 Mercedes-Benz #2638': "2022 Mercedes-Benz 2638",
    '2015 Ford Transit #6788': "2015 Ford Transit 6788",
    '2007 Ford Econoline E150 #2091': "2007 Ford Econoline E150 2091",
    'Annual Premium': "",
    'Combined Single Limit': "",
    'No. of Vehicles': "",
    'No. of Trailers': "",
    'Hired & Non_Owned Auto Liability': "Hired and Non_Owned Auto Liability",
    'Number of Autos': "",
    'Comprehensive - Owned Autos - Deductibles': "",
    'Collision - Owned Autos - Deductibles': "",
    'Physical Damage - Collision - Hired Autos - Deductibles': "",
    'Physical Damage - Comprehensive - Hired Autos -Deductibles': "",
    'Blanket Loss Payee': "",
    'Primary & Non-Contributory': "Primary and Non-Contributory",
    'Premium': "Premium Text",
    'Notes': "",

    # 'Garagekeepers Coverage Maximum': "",
    # 'Garagekeepers Coverage': "",
}

# add every key in dict in output excel sheet
df = pd.DataFrame(auto_review_dict.keys(), columns=['Auto Report Items'])

input_file = r"C:\Users\rrguest.RR-ITS\Documents\IMA Retail-09-09-2023\IMA Retail\Input Excels\Auto_preview.xlsx"
output_file = r'C:\Users\rrguest.RR-ITS\Documents\IMA Retail-09-09-2023\IMA Retail\Outphkut.xlsx'

df.to_excel(output_file, index=False, sheet_name="Auto")

workbook = load_workbook(output_file)
sheet = workbook['Auto']
sheet.freeze_panes = 'B2'
# Set fill color (sky blue) for the first column header
first_column_header = sheet.cell(row=1, column=1)
first_column_header.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
sheet.column_dimensions[first_column_header.column_letter].width = 60.00
workbook.save(output_file)

workbook = load_workbook(output_file)
sheet = workbook['Auto']


input_df = pd.read_excel(input_file)

for index, row in input_df.iterrows():
    file_name = row['File']
    col_index = sheet.max_column + 1
    column_header = sheet.cell(row=1, column=col_index, value=file_name)
    sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = 45.00
    column_header.fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    for key in auto_review_dict.keys():

        if key in input_df.columns:
            mapped_key = key
        else:
            mapped_key = auto_review_dict[key]


        sheet.cell(row=list(auto_review_dict.keys()).index(key) + 2, column=col_index, value=row[mapped_key])

# Enable text wrapping for all rows
for row in sheet.iter_rows(min_row=1, min_col=1, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')


workbook.save(output_file)

