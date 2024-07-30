from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def apply_conditional_formatting(output_file, final_df):
    # Load the workbook
    wb = load_workbook(output_file)
    ws = wb.active

    # Define fill colors
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    light_pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    # Loop through each triplet of columns in the final DataFrame
    for col in range(3, final_df.shape[1], 3):
        # print(col)

        reference_col_letter = ws.cell(row=1, column=col).column_letter
        # print(reference_col_letter)

        redrock_col_letter = ws.cell(row=1, column=col + 1).column_letter
        # print(redrock_col_letter)
        # Apply conditional formatting rules
        for row in range(2, final_df.shape[0] + 2):
            reference_cell = ws[f"{reference_col_letter}{row}"]
            # print(reference_cell.value)
            redrock_cell = ws[f"{redrock_col_letter}{row}"]
            # print(redrock_cell.value)

            # Assuming Attribute keyword in the excel sheet in the first column itself
            first_column_value = final_df.iloc[row - 2, 0]

            ref_value = reference_cell.value.strip() if reference_cell.value else ""

            redrock_value = redrock_cell.value.strip() if redrock_cell.value else ""

            if not ref_value and not redrock_value:
                # Both cells are empty, no highlighting
                continue
            elif not ref_value and redrock_value:
                # Reference cell is empty but RedRock cell is not empty, highlight in light yellow
                redrock_cell.fill = light_yellow_fill
            elif ref_value == redrock_value:
                # Both cells are populated and values match, highlight in light green
                redrock_cell.fill = light_green_fill
            else:

                redrock_cell.fill = light_pink_fill
    #
    # Save the workbook with formatting
    wb.save(output_file)
    print("All files saved in excel successfully")

