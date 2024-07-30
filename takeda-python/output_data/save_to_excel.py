from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from output_data.expected_values_with_regex import regex_expressions
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from output_data.fetch_xml import normalize_heading




def save_output_to_excel(main_samples, docugami_values, output_path):
    wb = Workbook()
    ws_main_samples = wb.active
    ws_main_samples.title = "Main Samples"

    # Define column headers
    headers_main_samples = ['Record ID', 'Record External ID', 'Record External Source', 'Study Name', 'Consent Type',
                            'Country',
                            'Site', 'Protocol Amendment Number', 'ICF Version Number', 'Effective Start Date',
                            'Sample Tracking Code',
                            'Attribute', 'Value']

    ws_main_samples.append(headers_main_samples)

    # Set column headers in bold
    for cell in ws_main_samples[1]:
        cell.font = Font(bold=True)

    sky_blue_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

    for i, entry in enumerate(main_samples, start=2):
        ws_main_samples.append(entry)

        if i % 4 == 2 or i % 4 == 0:  # Check if the row is the 2nd or 4th row of the group
            for cell in ws_main_samples[i]:
                cell.fill = sky_blue_fill

        attribute = entry[11]
        attribute = normalize_heading(attribute)
        if attribute in regex_expressions:
            options = list(regex_expressions[attribute].keys())
            cell = ws_main_samples.cell(row=i, column=13)
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', allow_blank=True)
            ws_main_samples.add_data_validation(dv)
            dv.add(cell)

    # Create Main CSV sheet
    ws_main_csv = wb.create_sheet(title="Main CSV")

    # Define column headers for Main CSV (same as Main Samples)
    headers_main_csv = headers_main_samples

    ws_main_csv.append(headers_main_csv)

    # Set column headers in bold for Main CSV
    for cell in ws_main_csv[1]:
        cell.font = Font(bold=True)

    for row in range(2, ws_main_samples.max_row + 1):
        for col in range(1, ws_main_samples.max_column + 1):
            cell_main_samples = ws_main_samples.cell(row=row, column=col)
            cell_main_csv = ws_main_csv.cell(row=row, column=col)
            if cell_main_samples.coordinate != 'M1':  # Skip the header row
                # Get the column letter for the current cell
                col_letter = get_column_letter(col)
                # Construct the formula referencing the cell in Main Samples
                formula = f'=IF(ISBLANK(\'Main Samples\'!{col_letter}{row}), "", \'Main Samples\'!{col_letter}{row})'
                cell_main_csv.value = formula

    # Docugami Values sheet
    ws_docugami_values = wb.create_sheet(title="Docugami Attributes and Answers")

    # Define column headers for Docugami Values
    docugami_values_headers = ['Attribute', 'Answer']

    # Append headers to Docugami Values sheet
    ws_docugami_values.append(docugami_values_headers)

    for cell in ws_docugami_values[1]:
        cell.font = Font(bold=True)

    for j, entry in enumerate(docugami_values, start=2):
        ws_docugami_values.append(entry)

        if j % 4 == 2 or j % 4 == 0:  # Check if the row is the 2nd or 4th row of the group
            for cell in ws_docugami_values[j]:
                cell.fill = sky_blue_fill

    wb.save(output_path)
